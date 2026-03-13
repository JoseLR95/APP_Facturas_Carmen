import os
import re
import io
import shutil
import tempfile
import zipfile
import streamlit as st
import pdfplumber
import openpyxl
from openpyxl import Workbook
from pypdf import PdfReader, PdfWriter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from collections import Counter

# ===========================================================================
# ESTILOS
# ===========================================================================

st.set_page_config(page_title="Gestión de Facturas", page_icon="📄", layout="centered")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;600&display=swap');

html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
.stApp { background: #0f0f0f; color: #e8e8e8; }
h1 {
    font-family: 'DM Mono', monospace !important;
    font-size: 1.6rem !important;
    color: #f0e040 !important;
    letter-spacing: -0.02em;
    margin-bottom: 0 !important;
}
.subtitle {
    font-family: 'DM Mono', monospace;
    font-size: 0.8rem;
    color: #555;
    margin-top: 0.2rem;
    margin-bottom: 2rem;
}
.block-label {
    font-family: 'DM Mono', monospace;
    font-size: 0.72rem;
    color: #f0e040;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    margin-bottom: 0.3rem;
}
.stat-box {
    background: #1a1a1a;
    border: 1px solid #2a2a2a;
    border-radius: 8px;
    padding: 1rem 1.2rem;
    text-align: center;
}
.stat-num {
    font-family: 'DM Mono', monospace;
    font-size: 2rem;
    font-weight: 500;
    color: #f0e040;
    line-height: 1;
}
.stat-label {
    font-size: 0.72rem;
    color: #666;
    margin-top: 0.3rem;
    text-transform: uppercase;
    letter-spacing: 0.05em;
}
.log-box {
    background: #111;
    border: 1px solid #222;
    border-radius: 8px;
    padding: 1rem 1.2rem;
    font-family: 'DM Mono', monospace;
    font-size: 0.72rem;
    color: #888;
    max-height: 320px;
    overflow-y: auto;
    line-height: 1.7;
}
.log-box .ok   { color: #6fcf97; }
.log-box .warn { color: #f2994a; }
.log-box .err  { color: #eb5757; }
.log-box .info { color: #56ccf2; }
section[data-testid="stFileUploadDropzone"] {
    background: #1a1a1a !important;
    border: 1px dashed #333 !important;
    border-radius: 8px !important;
}
div[data-testid="stDownloadButton"] button {
    background: #f0e040 !important;
    color: #0f0f0f !important;
    font-family: 'DM Mono', monospace !important;
    font-weight: 500 !important;
    border: none !important;
    border-radius: 6px !important;
    padding: 0.6rem 1.5rem !important;
    font-size: 0.85rem !important;
    width: 100%;
    margin-top: 1rem;
}
div[data-testid="stDownloadButton"] button:hover { background: #fff176 !important; }
.stProgress > div > div { background: #f0e040 !important; }
hr { border-color: #1e1e1e !important; margin: 1.5rem 0 !important; }
div[data-testid="stSelectbox"] > div {
    background: #1a1a1a !important;
    border: 1px solid #333 !important;
    border-radius: 8px !important;
    color: #e8e8e8 !important;
}
</style>
""", unsafe_allow_html=True)

# ===========================================================================
# AUTENTICACIÓN
# ===========================================================================

def check_password():
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if st.session_state.authenticated:
        return True
    st.markdown("<h1>📄 Gestión de Facturas</h1>", unsafe_allow_html=True)
    st.markdown("<div class='subtitle'>// Introduce la contraseña para acceder</div>", unsafe_allow_html=True)
    st.markdown("<div class='block-label'>Contraseña de acceso</div>", unsafe_allow_html=True)
    password = st.text_input("", type="password", placeholder="Introduce la contraseña", label_visibility="collapsed")
    if st.button("Acceder", use_container_width=True):
        if password.strip() == st.secrets["PASSWORD"].strip():
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("Contraseña incorrecta.")
    return False

if not check_password():
    st.stop()

# ===========================================================================
# SELECTOR DE MÓDULO
# ===========================================================================

st.markdown("<h1>📄 Gestión de Facturas</h1>", unsafe_allow_html=True)
st.markdown("<div class='subtitle'>// Selecciona el tipo de gestión</div>", unsafe_allow_html=True)
st.markdown("<div class='block-label'>Módulo</div>", unsafe_allow_html=True)
modulo = st.selectbox("", ["— Selecciona un módulo —", "Cheques", "Northgate", "Imputaciones"], label_visibility="collapsed")
st.markdown("<hr>", unsafe_allow_html=True)

# ===========================================================================
# MÓDULO 1: CHEQUES
# ===========================================================================

if modulo == "Cheques":

    st.markdown("<div class='subtitle'>// Renombrado de facturas cheque gourmet</div>", unsafe_allow_html=True)

    def cargar_excel_cheques(excel_bytes, fila_inicio=2):
        wb = openpyxl.load_workbook(excel_bytes, data_only=True)
        ws = wb.active
        filas = []
        for fila in ws.iter_rows(min_row=fila_inicio, values_only=True):
            codigo = fila[4]
            if codigo is None:
                continue
            codigo = str(codigo).strip()
            col_b  = str(fila[1]).strip() if fila[1] is not None else ""
            col_c  = str(fila[2]).strip() if fila[2] is not None else ""
            col_d  = str(fila[3]).strip() if fila[3] is not None else ""
            col_f  = str(fila[5]).strip() if fila[5] is not None else ""
            filas.append((codigo, col_b, col_c, col_d, col_f))
        conteo_excel     = Counter(f[0] for f in filas)
        duplicados_excel = {cod for cod, cnt in conteo_excel.items() if cnt > 1}
        mapeo = {}
        for codigo, col_b, col_c, col_d, col_f in filas:
            if codigo not in mapeo:
                mapeo[codigo] = (col_b, col_c, col_d, col_f)
        mapeo_duplicados = {}
        for codigo, col_b, col_c, col_d, col_f in filas:
            if codigo in duplicados_excel:
                if codigo not in mapeo_duplicados:
                    mapeo_duplicados[codigo] = []
                mapeo_duplicados[codigo].append((col_b, col_c, col_d, col_f))
        return mapeo, duplicados_excel, mapeo_duplicados

    def extraer_dnis_celda(col_f):
        return [d.strip().upper() for d in re.split(r'[\s/\-]+', col_f) if d.strip()]

    def normalizar_dni(dni):
        if re.match(r'^0\d{8}[A-Z]$', dni):
            return dni[1:]
        return dni

    def extraer_texto_cheques(ruta_pdf):
        texto = ""
        try:
            with pdfplumber.open(ruta_pdf) as pdf:
                for pagina in pdf.pages:
                    t = pagina.extract_text()
                    if t:
                        texto += t + "\n"
        except:
            pass
        return texto

    def es_anexo(texto):
        return "Anexo Factura Tarjeta Cheque Gourmet" in texto

    def extraer_codigo_cliente(texto):
        match = re.search(r'(\d{5})\s*N[oº°]\s*PROVEEDOR', texto)
        return match.group(1).strip() if match else None

    def extraer_num_factura_factura(texto):
        match = re.search(r'N[oº°]\s*PEDIDO[:\s]+(\w+)', texto)
        return match.group(1).strip() if match else None

    def extraer_num_factura_anexo(texto):
        lineas = texto.splitlines()
        for i, linea in enumerate(lineas):
            if "Nº FACTURA:" in linea and i > 0:
                return lineas[i - 1].strip()
        return None

    def extraer_dnis_anexo(texto):
        return [m.upper() for m in re.findall(r'\b(?:[0-9]{9}[A-Z]|[XYZ][0-9]{7}[A-Z])\b', texto)]

    def sanitizar_nombre(nombre):
        return re.sub(r'[\\/*?:"<>|]', "", nombre).strip()

    def procesar_cheques(pdf_files, excel_bytes, fila_inicio=2):
        logs  = []
        stats = {"unicos": 0, "repetidos_pdf": 0, "duplicados_resueltos": 0,
                 "duplicados_no_resueltos": 0, "no_encontrados": 0, "anexos": 0}

        def log(msg, tipo=""):
            logs.append((msg, tipo))

        try:
            mapeo, duplicados_excel, mapeo_duplicados = cargar_excel_cheques(excel_bytes, fila_inicio)
            log(f"Excel cargado — {len(mapeo)} códigos, {len(duplicados_excel)} duplicados", "info")
        except Exception as e:
            log(f"Error cargando Excel: {e}", "err")
            return None, logs, stats

        tmpdir = tempfile.mkdtemp()
        for f in pdf_files:
            ruta = os.path.join(tmpdir, f.name)
            with open(ruta, "wb") as out:
                out.write(f.read())

        pdfs = sorted(os.listdir(tmpdir))
        log(f"{len(pdfs)} PDFs recibidos", "info")

        facturas = {}
        anexos   = {}

        for nombre_pdf in pdfs:
            ruta  = os.path.join(tmpdir, nombre_pdf)
            texto = extraer_texto_cheques(ruta)
            if es_anexo(texto):
                num_factura = extraer_num_factura_anexo(texto)
                dnis        = [normalizar_dni(d) for d in extraer_dnis_anexo(texto)]
                if num_factura:
                    anexos[num_factura] = dnis
                log(f"[ANEXO] {nombre_pdf} → factura: {num_factura}, {len(dnis)} DNIs", "info")
                stats["anexos"] += 1
            else:
                codigo      = extraer_codigo_cliente(texto)
                num_factura = extraer_num_factura_factura(texto)
                facturas[nombre_pdf] = {"codigo": codigo, "num_factura": num_factura}
                log(f"[FACTURA] {nombre_pdf} → código: {codigo}", "")

        conteo_pdfs     = Counter(v["codigo"] for v in facturas.values() if v["codigo"])
        outdir          = tempfile.mkdtemp()
        contador_vistos = {}
        pendientes      = {}

        for nombre_pdf, datos in facturas.items():
            codigo      = datos["codigo"]
            num_factura = datos["num_factura"]
            if not codigo:
                log(f"⚠ {nombre_pdf} → sin código, ignorado", "warn")
                continue
            contador_vistos[codigo] = contador_vistos.get(codigo, 0) + 1
            n     = contador_vistos[codigo]
            total = conteo_pdfs[codigo]
            if codigo in duplicados_excel:
                pendientes[nombre_pdf] = {"codigo": codigo, "num_factura": num_factura, "n": n}
                continue
            if codigo not in mapeo:
                nuevo_nombre = sanitizar_nombre(f"{codigo}-No encontrado") + ".pdf"
                stats["no_encontrados"] += 1
            elif total == 1:
                col_b, col_c, col_d, _ = mapeo[codigo]
                partes = [p for p in [col_b, col_c, col_d] if p]
                nuevo_nombre = sanitizar_nombre("-".join(partes)) + ".pdf"
                stats["unicos"] += 1
            else:
                col_b, col_c, col_d, _ = mapeo[codigo]
                partes = [p for p in [col_b, col_c, col_d] if p]
                nuevo_nombre = sanitizar_nombre("-".join(partes)) + f"-{n}.pdf"
                stats["repetidos_pdf"] += 1
            shutil.copy2(os.path.join(tmpdir, nombre_pdf), os.path.join(outdir, nuevo_nombre))
            log(f"✓ {nombre_pdf} → {nuevo_nombre}", "ok")

        for nombre_pdf, datos in pendientes.items():
            codigo      = datos["codigo"]
            num_factura = datos["num_factura"]
            n           = datos["n"]
            nuevo_nombre = None
            log(f"[DUP] {nombre_pdf} → código: {codigo}, factura: {num_factura}", "warn")
            if num_factura and num_factura in anexos:
                dnis_anexo   = [normalizar_dni(d) for d in anexos[num_factura]]
                filas_codigo = mapeo_duplicados.get(codigo, [])
                for col_b, col_c, col_d, col_f in filas_codigo:
                    dnis_excel = [normalizar_dni(d) for d in extraer_dnis_celda(col_f)]
                    if any(dni in dnis_excel for dni in dnis_anexo):
                        partes = [p for p in [col_b, col_c, col_d] if p]
                        nuevo_nombre = sanitizar_nombre("-".join(partes)) + ".pdf"
                        stats["duplicados_resueltos"] += 1
                        log(f"  ✓ DNI coincide → {nuevo_nombre}", "ok")
                        break
            if not nuevo_nombre:
                nuevo_nombre = sanitizar_nombre(f"{codigo}-Duplicado-{n}") + ".pdf"
                stats["duplicados_no_resueltos"] += 1
                log(f"  → {nuevo_nombre}", "warn")
            shutil.copy2(os.path.join(tmpdir, nombre_pdf), os.path.join(outdir, nuevo_nombre))

        zip_path = os.path.join(tempfile.mkdtemp(), "facturas_renombradas.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in os.listdir(outdir):
                zf.write(os.path.join(outdir, f), f)
        return zip_path, logs, stats

    # UI Cheques
    st.markdown("<div class='block-label'>1 · Fichero Excel</div>", unsafe_allow_html=True)
    excel_file = st.file_uploader("", type=["xlsx"], key="excel_cheques", label_visibility="collapsed")
    st.markdown("<div class='block-label'>2 · PDFs de facturas y anexos</div>", unsafe_allow_html=True)
    pdf_files = st.file_uploader("", type=["pdf"], accept_multiple_files=True, key="pdfs_cheques", label_visibility="collapsed")
    fila_inicio = st.number_input("Fila de inicio (datos)", min_value=1, value=2, key="fila_cheques")
    st.markdown("<hr>", unsafe_allow_html=True)

    if st.button("▶ Procesar facturas Cheques", use_container_width=True):
        if not excel_file:
            st.error("Sube el fichero Excel.")
        elif not pdf_files:
            st.error("Sube al menos un PDF.")
        else:
            with st.spinner("Procesando..."):
                zip_path, logs, stats = procesar_cheques(pdf_files, excel_file, fila_inicio=int(fila_inicio))
            if zip_path:
                st.markdown("<hr>", unsafe_allow_html=True)
                total = sum(v for k, v in stats.items() if k != "anexos")
                c1, c2, c3, c4, c5 = st.columns(5)
                for col, label, key in [
                    (c1, "Únicos",       "unicos"),
                    (c2, "Repetidos",    "repetidos_pdf"),
                    (c3, "Dup. resuel.", "duplicados_resueltos"),
                    (c4, "Dup. pend.",   "duplicados_no_resueltos"),
                    (c5, "No encontr.",  "no_encontrados"),
                ]:
                    with col:
                        st.markdown(f"""
                        <div class='stat-box'>
                            <div class='stat-num'>{stats[key]}</div>
                            <div class='stat-label'>{label}</div>
                        </div>""", unsafe_allow_html=True)
                st.markdown(
                    f"<br><div style='text-align:center;font-family:DM Mono,monospace;font-size:0.8rem;color:#555'>"
                    f"TOTAL: <span style='color:#f0e040'>{total}</span> · "
                    f"ANEXOS IGNORADOS: <span style='color:#555'>{stats['anexos']}</span></div>",
                    unsafe_allow_html=True)
                st.markdown("<br><div class='block-label'>Log</div>", unsafe_allow_html=True)
                log_html = ""
                for msg, tipo in logs:
                    css = {"ok": "ok", "warn": "warn", "err": "err", "info": "info"}.get(tipo, "")
                    log_html += f"<div class='{css}'>{msg}</div>"
                st.markdown(f"<div class='log-box'>{log_html}</div>", unsafe_allow_html=True)
                with open(zip_path, "rb") as f:
                    st.download_button(
                        label="⬇ Descargar ZIP con facturas renombradas",
                        data=f.read(),
                        file_name="facturas_renombradas.zip",
                        mime="application/zip"
                    )

# ===========================================================================
# MÓDULO 2: NORTHGATE
# ===========================================================================

elif modulo == "Northgate":

    st.markdown("<div class='subtitle'>// Extracción de nº de pedido en facturas Northgate</div>", unsafe_allow_html=True)

    PATRON_CE_GUION   = r'\bCE-\d{4}-\d{10}\b'
    PATRON_CE_BARRA   = r'\bCE/\d{4}/\d{10}\b'
    PATRON_SC_NG      = r'\bSC-\d{6}-[A-Z]{2}_[A-Z]{3,6}-\d{2}\b'
    PATRON_MATRICULA  = r'\b\d{4}-[A-Z]{3}\b'
    PATRON_FACTURA_NG = r'N[oº°]\s*FACTURA\s+([A-Z]\d+)'

    def normalizar_matricula_ng(codigo):
        if re.match(r'^\d{4}-?[A-Z]{3}$', codigo):
            return codigo.replace("-", "")
        return codigo

    def normalizar_ce_ng(codigo):
        if re.match(r'^CE-\d{4}-\d{10}$', codigo):
            partes = codigo.split("-")
            return f"{partes[0]}/{partes[1]}/{partes[2]}"
        if re.match(r'^CE/\d{4}/\d{10}$', codigo):
            partes = codigo.split("/")
            return f"{partes[0]}-{partes[1]}-{partes[2]}"
        return codigo

    def cargar_excel_ng(excel_bytes, fila_inicio=2):
        wb = openpyxl.load_workbook(excel_bytes, data_only=True)
        ws = wb.active
        mapeo_a, mapeo_b, mapeo_c = {}, {}, {}
        for fila in ws.iter_rows(min_row=fila_inicio, values_only=True):
            col_a = str(fila[0]).strip() if fila[0] is not None else ""
            col_b = str(fila[1]).strip() if fila[1] is not None else ""
            col_c = str(fila[2]).strip() if fila[2] is not None else ""
            col_d = str(fila[3]).strip() if fila[3] is not None else ""
            if col_a:
                mapeo_a[normalizar_matricula_ng(col_a)] = col_d
            if col_b:
                mapeo_b[col_b] = col_d
                mapeo_b[normalizar_ce_ng(col_b)] = col_d
            if col_c:
                mapeo_c[col_c] = col_d
        return mapeo_a, mapeo_b, mapeo_c

    def extraer_texto_ng(ruta_pdf):
        texto = ""
        try:
            with pdfplumber.open(ruta_pdf) as pdf:
                for pagina in pdf.pages:
                    t = pagina.extract_text()
                    if t:
                        texto += t + "\n"
        except:
            pass
        return texto

    def extraer_numero_factura_ng(texto):
        match = re.search(PATRON_FACTURA_NG, texto)
        return match.group(1).strip() if match else None

    def extraer_codigos_linea_ng(linea):
        codigos = []
        codigos += re.findall(PATRON_CE_GUION, linea)
        codigos += re.findall(PATRON_CE_BARRA, linea)
        codigos += re.findall(PATRON_SC_NG, linea)
        codigos += re.findall(PATRON_MATRICULA, linea)
        return list(dict.fromkeys(codigos))

    def extraer_lineas_ng(texto):
        lineas = []
        for linea in texto.splitlines():
            if "FACTURACI" in linea.upper() and "CONTRATO" in linea.upper():
                lineas.append(linea)
            elif re.search(PATRON_SC_NG, linea):
                lineas.append(linea)
            elif re.search(PATRON_CE_BARRA, linea) or re.search(PATRON_CE_GUION, linea):
                # Líneas con CE pero sin SC (ej: RTT - NOMBRE CE/xxxx/... matrícula)
                lineas.append(linea)
        return lineas

    def buscar_valor_d_ng(codigo, mapeo_a, mapeo_b, mapeo_c):
        if codigo in mapeo_b: return mapeo_b[codigo]
        if normalizar_ce_ng(codigo) in mapeo_b: return mapeo_b[normalizar_ce_ng(codigo)]
        if codigo in mapeo_c: return mapeo_c[codigo]
        if normalizar_matricula_ng(codigo) in mapeo_a: return mapeo_a[normalizar_matricula_ng(codigo)]
        return None

    def buscar_po_en_linea_ng(linea, mapeo_a, mapeo_b, mapeo_c):
        codigos = extraer_codigos_linea_ng(linea)
        for codigo in codigos:
            valor_d = buscar_valor_d_ng(codigo, mapeo_a, mapeo_b, mapeo_c)
            if valor_d:
                # Preferir SC, si no matrícula, si no CE, si no el primero disponible
                codigo_rep = (
                    next((c for c in codigos if re.match(PATRON_SC_NG, c)), None) or
                    next((c for c in codigos if re.match(PATRON_MATRICULA, c)), None) or
                    next((c for c in codigos if re.match(PATRON_CE_BARRA, c) or re.match(PATRON_CE_GUION, c)), None) or
                    codigos[0]
                )
                return codigo_rep, valor_d
        codigo_rep = (
            next((c for c in codigos if re.match(PATRON_SC_NG, c)), None) or
            next((c for c in codigos if re.match(PATRON_MATRICULA, c)), None) or
            next((c for c in codigos if re.match(PATRON_CE_BARRA, c) or re.match(PATRON_CE_GUION, c)), None) or
            (codigos[0] if codigos else "")
        )
        return codigo_rep, None

    def crear_pagina_resumen_ng(num_factura, filas):
        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(buffer, pagesize=A4,
                                   leftMargin=40, rightMargin=40,
                                   topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        story  = []
        story.append(Paragraph(f"<b>Nº Pedido — Factura {num_factura}</b>", styles["Title"]))
        story.append(Spacer(1, 20))
        datos_tabla = [["Código SC", "Nº Pedido"]]
        for codigo, valor_d in filas:
            datos_tabla.append([codigo, valor_d])
        tabla = Table(datos_tabla, colWidths=[250, 250])
        tabla.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0),  colors.HexColor("#1a1a1a")),
            ("TEXTCOLOR",      (0, 0), (-1, 0),  colors.HexColor("#f0e040")),
            ("FONTNAME",       (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",       (0, 0), (-1, 0),  10),
            ("ALIGN",          (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME",       (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",       (0, 1), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f9f9f9"), colors.white]),
            ("GRID",           (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ("TOPPADDING",     (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 6),
            ("LEFTPADDING",    (0, 0), (-1, -1), 8),
        ]))
        story.append(tabla)
        doc.build(story)
        buffer.seek(0)
        return buffer

    def añadir_pagina_ng(ruta_origen, ruta_destino, num_factura, filas):
        pagina_extra    = crear_pagina_resumen_ng(num_factura, filas)
        reader_original = PdfReader(ruta_origen)
        reader_extra    = PdfReader(pagina_extra)
        writer          = PdfWriter()
        for pagina in reader_original.pages:
            writer.add_page(pagina)
        for pagina in reader_extra.pages:
            writer.add_page(pagina)
        with open(ruta_destino, "wb") as f:
            writer.write(f)

    def procesar_northgate(pdf_files, excel_bytes, fila_inicio=2):
        logs  = []
        stats = {"gestionadas": 0, "no_gestionadas": 0}

        def log(msg, tipo=""):
            logs.append((msg, tipo))

        try:
            mapeo_a, mapeo_b, mapeo_c = cargar_excel_ng(excel_bytes, fila_inicio)
            log(f"Excel cargado correctamente", "info")
        except Exception as e:
            log(f"Error cargando Excel: {e}", "err")
            return None, logs, stats

        tmpdir = tempfile.mkdtemp()
        for f in pdf_files:
            ruta = os.path.join(tmpdir, f.name)
            with open(ruta, "wb") as out:
                out.write(f.read())

        pdfs = sorted([f for f in os.listdir(tmpdir) if f.lower().endswith(".pdf")])
        log(f"{len(pdfs)} PDFs recibidos", "info")

        outdir_gest   = tempfile.mkdtemp()
        outdir_nogest = tempfile.mkdtemp()
        resultados    = {}
        no_gestionadas = []

        for nombre_pdf in pdfs:
            ruta  = os.path.join(tmpdir, nombre_pdf)
            texto = extraer_texto_ng(ruta)
            num_factura        = extraer_numero_factura_ng(texto)
            lineas_facturacion = extraer_lineas_ng(texto)

            if not num_factura:
                num_factura = nombre_pdf.replace(".PDF", "").replace(".pdf", "")
                log(f"⚠ {nombre_pdf} → nº factura no detectado", "warn")

            log(f"[PDF] {nombre_pdf} → {num_factura}, {len(lineas_facturacion)} línea(s)", "")

            filas = []
            for linea in lineas_facturacion:
                codigo_rep, valor_po = buscar_po_en_linea_ng(linea, mapeo_a, mapeo_b, mapeo_c)
                if valor_po:
                    filas.append((codigo_rep, valor_po))
                    log(f"  ✓ {codigo_rep} → {valor_po}", "ok")
                else:
                    log(f"  ❌ {codigo_rep} → NO ENCONTRADO", "warn")

            if filas:
                resultados[num_factura] = filas
                stats["gestionadas"] += 1
                nombre_nuevo = nombre_pdf.replace(".PDF", "_modificado.pdf").replace(".pdf", "_modificado.pdf")
                try:
                    añadir_pagina_ng(ruta, os.path.join(outdir_gest, nombre_nuevo), num_factura, filas)
                    log(f"  → PDF guardado: {nombre_nuevo}", "ok")
                except Exception as e:
                    log(f"  [ERROR] PDF: {e}", "err")
            else:
                no_gestionadas.append(num_factura)
                stats["no_gestionadas"] += 1
                shutil.copy2(ruta, os.path.join(outdir_nogest, nombre_pdf))
                log(f"  → Sin PO, movido a No gestionadas", "warn")

        # Crear Excel
        wb = Workbook()
        wb.remove(wb.active)
        for num_factura, filas in sorted(resultados.items()):
            ws = wb.create_sheet(title=str(num_factura)[:31])
            ws.append(["Código SC", "Nº Pedido"])
            for codigo, valor_d in filas:
                ws.append([codigo, valor_d])
        if no_gestionadas:
            ws_ng = wb.create_sheet(title="No gestionadas")
            ws_ng.append(["Nº Factura"])
            for nf in sorted(no_gestionadas):
                ws_ng.append([nf])
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Crear ZIP único con PDFs + carpeta No gestionadas + Excel
        zip_final = io.BytesIO()
        with zipfile.ZipFile(zip_final, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in os.listdir(outdir_gest):
                zf.write(os.path.join(outdir_gest, f), f)
            for f in os.listdir(outdir_nogest):
                zf.write(os.path.join(outdir_nogest, f), os.path.join("No gestionadas", f))
            zf.writestr("resultado_northgate.xlsx", excel_buffer.getvalue())
        zip_final.seek(0)

        return zip_final, logs, stats

    # UI Northgate
    st.markdown("<div class='block-label'>1 · Fichero Excel de referencias</div>", unsafe_allow_html=True)
    excel_ng = st.file_uploader("", type=["xlsx"], key="excel_ng", label_visibility="collapsed")
    st.markdown("<div class='block-label'>2 · PDFs de facturas</div>", unsafe_allow_html=True)
    pdfs_ng = st.file_uploader("", type=["pdf"], accept_multiple_files=True, key="pdfs_ng", label_visibility="collapsed")
    fila_ng = st.number_input("Fila de inicio (datos)", min_value=1, value=2, key="fila_ng")
    st.markdown("<hr>", unsafe_allow_html=True)

    if st.button("▶ Procesar facturas Northgate", use_container_width=True):
        if not excel_ng:
            st.error("Sube el fichero Excel.")
        elif not pdfs_ng:
            st.error("Sube al menos un PDF.")
        else:
            with st.spinner("Procesando..."):
                zip_final, logs, stats = procesar_northgate(pdfs_ng, excel_ng, fila_inicio=int(fila_ng))
            if zip_final:
                st.markdown("<hr>", unsafe_allow_html=True)
                c1, c2 = st.columns(2)
                with c1:
                    st.markdown(f"""
                    <div class='stat-box'>
                        <div class='stat-num'>{stats['gestionadas']}</div>
                        <div class='stat-label'>Gestionadas</div>
                    </div>""", unsafe_allow_html=True)
                with c2:
                    st.markdown(f"""
                    <div class='stat-box'>
                        <div class='stat-num'>{stats['no_gestionadas']}</div>
                        <div class='stat-label'>No gestionadas</div>
                    </div>""", unsafe_allow_html=True)

                st.markdown("<br><div class='block-label'>Log</div>", unsafe_allow_html=True)
                log_html = ""
                for msg, tipo in logs:
                    css = {"ok": "ok", "warn": "warn", "err": "err", "info": "info"}.get(tipo, "")
                    log_html += f"<div class='{css}'>{msg}</div>"
                st.markdown(f"<div class='log-box'>{log_html}</div>", unsafe_allow_html=True)

                st.download_button(
                    label="⬇ Descargar ZIP completo (PDFs + Excel)",
                    data=zip_final,
                    file_name="northgate_completo.zip",
                    mime="application/zip"
                )


# ===========================================================================
# MÓDULO 3: PO LOOKUP
# ===========================================================================

elif modulo == "Imputaciones":

    from openpyxl.styles import Border, Side
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm

    st.markdown("<div class='subtitle'>// Búsqueda de órdenes de compra en facturas PDF</div>", unsafe_allow_html=True)

    FILA_DATOS_PO = 17
    COL_A_PO=1; COL_L_PO=12; COL_M_PO=13; COL_R_PO=18; COL_T_PO=20; COL_V_PO=22; COL_AA_PO=27

    RE_PO_PAT = re.compile(r'\bPO[-/\s]?\d{6}[-/\s]?ES[_\-/\s][A-Z]{2,6}(?:[_\-][A-Z])?[-/\s]\d{2}\b')
    RE_PO_STILL_PAT = re.compile(r'PEDIDO\s*PO-(\d{6}-ES[-_][A-Z]{2,6}(?:[-_][A-Z])?-\d{2})', re.IGNORECASE)
    RE_SC_PAT = re.compile(r'\bSC-\d{6}-[A-Z]{2}_[A-Z]{2,6}(?:_[A-Z])?-\d{2}\b')
    RE_MAT_PAT = re.compile(r'\b\d{4}-[A-Z]{3}\b')

    def po3_norm(t):
        return re.sub(r'[-/_\s]', '', str(t)).upper()

    def po3_v(v):
        return str(v).strip() if v is not None else ""

    @st.cache_data(show_spinner="Cargando Excel interno…")
    def po3_cargar_excel(excel_bytes):
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
        ws = wb.active
        mp={};ms={};mi={}
        for fila in ws.iter_rows(min_row=FILA_DATOS_PO, values_only=True):
            po_raw=fila[COL_A_PO-1]
            if not po_raw: continue
            po_orig=po3_v(po_raw)
            t_cod=po3_v(fila[COL_T_PO-1]).split()[0] if po3_v(fila[COL_T_PO-1]) else ""
            col_b=" — ".join(p for p in [po3_v(fila[COL_R_PO-1]),t_cod,po3_v(fila[COL_V_PO-1])] if p)
            costo=po3_v(fila[COL_M_PO-1])
            datos=(po_orig,col_b,costo)
            mp[po3_norm(po_orig)]=datos
            aa=po3_v(fila[COL_AA_PO-1])
            if aa: ms[aa]=datos
            id2=po3_v(fila[COL_L_PO-1])
            if id2 and ' ' not in id2 and 4<=len(id2)<=30:
                mi[id2.upper()]=datos
                sg=id2.replace('-','').upper()
                if sg!=id2.upper(): mi[sg]=datos
        wb.close()
        return mp,ms,mi

    def po3_buscar(linea,mp,ms,mi):
        for m in RE_PO_PAT.finditer(linea):
            c=po3_norm(m.group())
            if c in mp: return ('PO',m.group(),mp[c])
        for m in RE_PO_STILL_PAT.finditer(linea):
            c=po3_norm('PO-'+m.group(1))
            if c in mp: return ('PO','PO-'+m.group(1),mp[c])
        for m in RE_SC_PAT.finditer(linea):
            if m.group() in ms: return ('SC',m.group(),ms[m.group()])
        for m in RE_MAT_PAT.finditer(linea):
            for c in [m.group().upper(),m.group().replace('-','').upper()]:
                if c in mi: return ('MAT',m.group(),mi[c])
        return None

    def po3_procesar_pdf(pdf_bytes,mp,ms,mi):
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            texto="\n".join(p.extract_text() or "" for p in pdf.pages)
        num=""
        for pat in [r'N[oº°\s]*FACTURA[:\s]+([A-Z0-9/\-]+)',r'Nº\s*FACTURA[:\s]+([A-Z0-9/\-]+)',
                    r'Factura\s+(?:nº|N[º°])[:\s]*(\S+)',r'Factura\s+MN\s*/\s*(\S+)',r'INVOICE[:\s]+([A-Z0-9\-/]+)']:
            m=re.search(pat,texto,re.IGNORECASE)
            if m: num=m.group(1).strip().rstrip('/'); break
        filas=[]; vistas=set()
        for linea in texto.splitlines():
            res=po3_buscar(linea.strip(),mp,ms,mi)
            if not res: continue
            tipo,ref,(po_orig,col_b,costo)=res
            c=po3_norm(po_orig)
            if c not in vistas:
                vistas.add(c)
                filas.append({"tipo":tipo,"ref":ref,"po":po_orig,"col_b":col_b,"costo":costo})
        return num,filas

    def po3_crear_excel(nombre_pdf,filas):
        from openpyxl.styles import Border, Side
        BORDE=Border(left=Side(style='thin',color='AAAAAA'),right=Side(style='thin',color='AAAAAA'),
                     top=Side(style='thin',color='AAAAAA'),bottom=Side(style='thin',color='AAAAAA'))
        wb=Workbook(); ws=wb.active; ws.title="Coincidencias"
        ws.append(["Orden de compra (PO)","Centro de costos — Proyecto — Client ID","Costo por unidad"])
        for cell in ws[1]:
            cell.fill=PatternFill("solid",fgColor="1A1A1A"); cell.font=Font(bold=True,color="F0E040",size=10)
            cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=BORDE
        ws.row_dimensions[1].height=22
        fp=PatternFill("solid",fgColor="F2F2F2"); fi=PatternFill("solid",fgColor="FFFFFF")
        for idx,f in enumerate(filas,1):
            ws.append([f["po"],f["col_b"],f["costo"]])
            for cell in ws[ws.max_row]:
                cell.fill=fp if idx%2==0 else fi; cell.border=BORDE; cell.alignment=Alignment(vertical="center")
        ws.column_dimensions['A'].width=30; ws.column_dimensions['B'].width=65; ws.column_dimensions['C'].width=18
        buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

    def po3_crear_resumen(num,filas):
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        buf=io.BytesIO()
        doc=SimpleDocTemplate(buf,pagesize=A4,leftMargin=1.5*cm,rightMargin=1.5*cm,topMargin=1.5*cm,bottomMargin=1.5*cm)
        styles=getSampleStyleSheet()
        amarillo=colors.HexColor("#f0e040"); oscuro=colors.HexColor("#1a1a1a")
        gris=colors.HexColor("#2a2a2a"); gris2=colors.HexColor("#222222"); borde=colors.HexColor("#555555")
        ts=ParagraphStyle('t',parent=styles['Heading2'],textColor=amarillo,backColor=oscuro,spaceAfter=10,spaceBefore=0,leftIndent=6)
        story=[Paragraph(f"Coincidencias — Factura {num or '(sin número)'}", ts), Spacer(1,0.3*cm)]
        ancho=A4[0]-3*cm
        datos=[["PO","Centro de costos — Proyecto — Client ID"]]
        for f in filas: datos.append([f["po"],f["col_b"]])
        tabla=Table(datos,colWidths=[0.35*ancho,0.65*ancho],repeatRows=1)
        tabla.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),oscuro),("TEXTCOLOR",(0,0),(-1,0),amarillo),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),8),
            ("FONTSIZE",(0,1),(-1,-1),7),("TEXTCOLOR",(0,1),(-1,-1),colors.white),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[gris,gris2]),("GRID",(0,0),(-1,-1),0.4,borde),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ]))
        story.append(tabla); doc.build(story); buf.seek(0); return buf.read()

    def po3_añadir_pagina(orig,pagina):
        writer=PdfWriter()
        for p in PdfReader(io.BytesIO(orig)).pages: writer.add_page(p)
        for p in PdfReader(io.BytesIO(pagina)).pages: writer.add_page(p)
        buf=io.BytesIO(); writer.write(buf); buf.seek(0); return buf.read()

    # ── UI ──────────────────────────────────────────────────────────────────
    st.markdown("<div class='block-label'>1 · Excel interno de órdenes de compra</div>", unsafe_allow_html=True)
    excel_po=st.file_uploader("",type=["xlsx"],key="excel_po",label_visibility="collapsed")
    st.markdown("<div class='block-label'>2 · PDFs de facturas</div>", unsafe_allow_html=True)
    pdfs_po=st.file_uploader("",type=["pdf"],accept_multiple_files=True,key="pdfs_po",label_visibility="collapsed")

    if excel_po and pdfs_po:
        mp,ms,mi=po3_cargar_excel(excel_po.read())
        st.success(f"✅ Excel cargado — **{len(mp):,}** POs · **{len(ms):,}** SCs · **{len(mi):,}** IDs")

        if st.button("Procesar",key="btn_po",use_container_width=True):
            log_lines=[]; stats={"total":0,"con_match":0,"sin_match":0,"por_PO":0,"por_SC":0,"por_MAT":0}
            zip_buf=io.BytesIO()
            prog=st.progress(0,text="Procesando…")
            with zipfile.ZipFile(zip_buf,"w",zipfile.ZIP_DEFLATED) as zf:
                for idx,pdf_file in enumerate(pdfs_po):
                    prog.progress(idx/len(pdfs_po),text=f"Procesando {pdf_file.name}…")
                    stats["total"]+=1
                    try:
                        pdf_bytes=pdf_file.read()
                        num,filas=po3_procesar_pdf(pdf_bytes,mp,ms,mi)
                        if not filas:
                            stats["sin_match"]+=1
                            log_lines.append(f'<span class="warn">⚠ {pdf_file.name} — sin coincidencias</span>')
                            continue
                        stats["con_match"]+=1
                        for f in filas:
                            stats[f"por_{f['tipo']}"]+=1
                            log_lines.append(f'<span class="ok">✓ {pdf_file.name} [{f["tipo"]}] {f["ref"]} → {f["po"]}</span>')
                        stem=pdf_file.name.rsplit('.',1)[0]
                        zf.writestr(stem+".xlsx", po3_crear_excel(pdf_file.name,filas))
                        zf.writestr(pdf_file.name, po3_añadir_pagina(pdf_bytes, po3_crear_resumen(num,filas)))
                    except Exception as e:
                        stats["sin_match"]+=1
                        log_lines.append(f'<span class="err">✗ {pdf_file.name} — Error: {e}</span>')
            prog.progress(1.0,text="¡Completado!")
            c1,c2,c3,c4,c5=st.columns(5)
            for col,label,key in [(c1,"PDFs total","total"),(c2,"Con match","con_match"),
                                   (c3,"Sin match","sin_match"),(c4,"Match PO/SC","por_PO"),(c5,"Match MAT","por_MAT")]:
                col.markdown(f"<div class='stat-box'><div class='stat-num'>{stats[key]}</div><div class='stat-label'>{label}</div></div>",unsafe_allow_html=True)
            st.markdown("<div class='log-box'>"+"<br>".join(log_lines)+"</div>",unsafe_allow_html=True)
            zip_buf.seek(0)
            st.download_button("⬇️ Descargar ZIP (PDFs + Excel)",data=zip_buf.read(),
                file_name="Imputaciones.zip",mime="application/zip",use_container_width=True)

# ===========================================================================
# PANTALLA INICIAL
# ===========================================================================

else:
    st.markdown("""
    <div style='text-align:center; padding: 3rem 0; color: #333;'>
        <div style='font-family: DM Mono, monospace; font-size: 0.85rem;'>
            ← Selecciona un módulo para comenzar
        </div>
    </div>
    """, unsafe_allow_html=True)
